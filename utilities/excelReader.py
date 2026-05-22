import openpyxl
import os


def get_excel_data(sheet_name):
    base_path = os.path.dirname(os.path.abspath(__file__))
    # Assuming your file is named purchaseData.xlsx inside utilities or a data folder
    file_path = os.path.join(base_path, "ProductData.xls")

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]

    max_row = sheet.max_row
    max_col = sheet.max_column

    test_data = []

    # Start from row 2 to skip headers
    for i in range(2, max_row + 1):
        row_data = []
        for j in range(1, max_col + 1):
            row_data.append(sheet.cell(row=i, column=j).value)
        test_data.append(tuple(row_data))

    return test_data